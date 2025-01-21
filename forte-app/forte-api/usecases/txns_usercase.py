import os
import pandas as pd

from flask import send_file
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

from adapters import db_adapters
from repository import b3_repository


class TnxsUsecase:

    def __init__(self, db):
        self.db = db

    def save_transaction(self, txn):
        txn_id = db_adapters.create_transaction(self.db, txn)
        return txn_id

    def update_transaction(self, transaction_id, data):
        updated = db_adapters.update_transaction(self.db, transaction_id, data)
        return updated

    def get_transactions_by_customer_id(self, customer_id):
        filter_by = {"customer_id": customer_id}
    
        txns = db_adapters.get_transactions(self.db, filter_by)
        return txns

    def get_all_transactions(self):
        txn = db_adapters.get_transactions(self.db, {})
        return txn

    def deactived_transaction(self, transaction_id):
        deactived = db_adapters.deactived_transaction(self.db, transaction_id)
        return deactived

    def get_transactions_resume(self, start_date, end_date):
        txns_list = db_adapters.get_transactions_by_date_range(self.db, start_date, end_date)

        df = pd.DataFrame(txns_list)

        txns_by_category = (
            df.groupby(["category", "type"])
                .agg({"amount": "sum"})
        )

        df['date'] = df['date'].dt.strftime('%d-%m-%Y')
        txns_by_customer = (
            df.groupby(["date", "type"])
                .agg({"amount": ["mean", "sum"]})["amount"]
        )
        
        txns_resume = {
            "by_category": df_to_dict(txns_by_category),
            "by_customer": df_to_dict(txns_by_customer)
        }

        return txns_resume


class TnxsResumeXLSX:
    def get_file(self, resume_by_category, txns_by_customer):

        wb = Workbook(write_only=True)

        # Save xlsx to transactions resume by categoy
        ws = wb.create_sheet(title="by category")

        header = ('category', 'incomes', 'expenses')  
        ws.append(header)

        for key, value in dict(resume_by_category).items():
            income = value.get('income', {}).get('amount', 0)
            expense = value.get('expense', {}).get('amount', 0)

            ws.append((key, income, expense))

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Resume by Category (income and expenses)"
        chart1.y_axis.title = 'Amount summatory'
        chart1.x_axis.title = 'Transactions'

        data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
        chart1.add_data(data, titles_from_data=True)

        cats = Reference(ws, min_col=1, min_row=2, max_row=7)
        chart1.set_categories(cats)
        
        chart1.shape = 4

        ws.add_chart(chart1, "G2") 

        # Seve xlsx transactions resume xlsx by customers
        ws1 = wb.create_sheet(title="by customers")
        
        header = ('date', 'incomes', 'expenses')  
        ws1.append(header)

        for key, value in dict(txns_by_customer).items():
            income = value.get('income', {}).get('sum', 0)
            expense = value.get('expense', {}).get('mean', 0)

            ws1.append((key, income, expense))

        chart2 = LineChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "Transactions resume by cutomers."
        chart2.y_axis.title = 'Amount'
        chart2.x_axis.title = 'Transactions'

        data = Reference(ws1, min_col=2, min_row=1, max_row=7, max_col=3)
        chart2.add_data(data, titles_from_data=True)

        cats = Reference(ws1, min_col=1, min_row=2, max_row=7)
        chart2.set_categories(cats)
        
        chart2.shape = 4

        ws1.add_chart(chart2, "G2") 

        keys = tuple(txns_by_customer.keys())
        xlsx_name = f"/txns_resume_{keys[0]}_{keys[-1]}.xlsx"

        if os.environ.get("AWS_EXECUTION_ENV"):
            xlsx_directory = f"/tmp{xlsx_name}"
        else:
            xlsx_directory = os.path.abspath(os.curdir) + xlsx_name

        wb.save(xlsx_directory)

        xlsx_file = send_file(
            xlsx_directory,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True
        )

        return xlsx_file
    

    def upload_file(self, xlsx_file, content_type):

        boto3 = b3_repository.Boto3()
        presigned_url = boto3.upload_file(xlsx_file, content_type)

        return presigned_url


def df_to_dict(df):
    if df.ndim == 1:
        return df.to_dict()

    ret = {}
    for key in df.index.get_level_values(0):
        sub_df = df.xs(key)
        ret[key] = df_to_dict(sub_df)
    return ret
